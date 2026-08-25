import hashlib
import io
import time
from copy import deepcopy
from decimal import Decimal

import pdfplumber
from openpyxl import load_workbook

from app.adapters.cl_din import WATERMARK, din_payload, render_din_pdf
from app.adapters.xlsx import TEMPLATE_PATH, WORKBOOK_NOTICE, build_workbook
from app.engine.reconcile import reconcile


def _state(scenario_b, chile_cfg, falabella_cfg, demo_fx):
    result = reconcile(scenario_b, chile_cfg, falabella_cfg, *demo_fx)
    return {
        "dispatch": {
            "despacho_no": "700612",
            "referencia": "54415CLFA/26J21-3",
            "status": "review",
            "regime": "import_for_consumption",
            "jurisdiction_config_hash": "c" * 64,
            "client_config_hash": "e" * 64,
            "din_acceptance_date": "2026-08-18",
        },
        "documents": [],
        "calculation": result,
        "audit": [],
        "calculation_run": {
            "input_hash": "d" * 64,
            "engine_version": "v1",
            "created_at": "2026-08-19T12:00:00+00:00",
        },
    }


def test_excel_has_required_sheets_and_totals(scenario_b, chile_cfg, falabella_cfg, demo_fx):
    template_hash_before = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()
    content = build_workbook(_state(scenario_b, chile_cfg, falabella_cfg, demo_fx))
    assert hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest() == template_hash_before
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    required = [
        "Resumen",
        "Documentos",
        "Extracciones",
        "Validaciones",
        "Prorrateo",
        "Tributos",
        "Vista declaración",
        "Vista costo",
        "Trazabilidad",
    ]
    assert all(name in workbook.sheetnames for name in required)
    assert workbook.sheetnames[:2] == ["Prorrateo General", "Prorrateo resumen"]
    values = {row[0].value: row[1].value for row in workbook["Resumen"].iter_rows(min_row=2)}
    assert values["Aviso"] == WORKBOOK_NOTICE
    assert values["Tributos USD"] == "13719.25"
    assert values["Costo puesto USD"] == "69139.20"
    assert values["Pago estimado CLP"] == "13217811"
    assert values["SHA-256 plantilla"] == template_hash_before
    assert values["SHA-256 configuración"] == "c" * 64
    assert values["SHA-256 cálculo"] == "d" * 64
    master = workbook["Prorrateo General"]
    assert master["F1"].value == "Prima póliza cliente"
    assert sum(Decimal(str(master[f"F{row}"].value)) for row in range(2, 5)) == Decimal("36.41")
    assert master["H2"].value == "=ROUND(G2*O2,2)"
    trace_values = [row[2].value for row in workbook["Trazabilidad"].iter_rows(min_row=2)]
    assert any(template_hash_before in str(value) for value in trace_values)


def test_din_pdf_is_readable_and_watermarked(scenario_b, chile_cfg, falabella_cfg, demo_fx):
    content = render_din_pdf(_state(scenario_b, chile_cfg, falabella_cfg, demo_fx))
    assert content.startswith(b"%PDF")
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        assert len(pdf.pages) == 3
    assert WATERMARK in text
    assert "13,719.25" not in text  # Report uses exact unlocalized value.
    assert "2539.02" in text


def test_din_json_is_one_declaration_per_invoice(scenario_b, chile_cfg, falabella_cfg, demo_fx):
    payload = din_payload(_state(scenario_b, chile_cfg, falabella_cfg, demo_fx))
    assert [item["invoice"] for item in payload] == [
        "BN26010512",
        "BN26010513",
        "BN26010514",
    ]
    assert all(len(item["lines"]) == 1 for item in payload)
    assert payload[2]["declaration_view"]["total_payable"] == "2539.02"


def test_exports_support_100_invoices_without_generation_pagination(
    scenario_b, chile_cfg, falabella_cfg, demo_fx
):
    state = _state(scenario_b, chile_cfg, falabella_cfg, demo_fx)
    template = state["calculation"]["lines"][0]
    lines = []
    for index in range(100):
        line = deepcopy(template)
        line["invoice"] = f"SCALE-{index + 1:03d}"
        lines.append(line)
    state["calculation"]["lines"] = lines
    started = time.perf_counter()
    payload = din_payload(state)
    pdf_content = render_din_pdf(state)
    workbook_content = build_workbook(state)
    elapsed = time.perf_counter() - started

    assert len(payload) == 100
    with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
        assert len(pdf.pages) == 100
    workbook = load_workbook(io.BytesIO(workbook_content), data_only=False)
    assert workbook["Prorrateo General"]["N101"].value == "SCALE-100"
    assert workbook["Prorrateo General"]["C102"].value == "=SUM(C2:C101)"
    assert elapsed < 300


def test_excel_uses_one_master_and_summary_row_per_multiline_invoice(
    scenario_b, chile_cfg, falabella_cfg, demo_fx
):
    invoice = scenario_b.invoices[0]
    original = invoice.lines[0]
    first = original.model_copy(deep=True)
    second = original.model_copy(deep=True)
    for line in (first, second):
        line.quantity.value = original.quantity.value / 2
        line.line_total.value = original.line_total.value / 2
    invoice.lines = [first, second]

    state = _state(scenario_b, chile_cfg, falabella_cfg, demo_fx)
    workbook = load_workbook(io.BytesIO(build_workbook(state)), data_only=False)

    assert [workbook["Prorrateo General"][f"N{row}"].value for row in range(2, 5)] == [
        "BN26010512",
        "BN26010513",
        "BN26010514",
    ]
    assert workbook["Vista declaración"].max_row == 4
    assert workbook["Vista costo"].max_row == 4
