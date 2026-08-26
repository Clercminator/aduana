from __future__ import annotations

import hashlib
import io
from copy import copy
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "08213F"
RED = "C81E1E"
TEMPLATE_PATH = Path(__file__).parents[2] / "PRORRATEO MASTER.xlsx"
WORKBOOK_NOTICE = "RESULTADO DE DEMOSTRACIÓN — VALIDACIÓN ADUANERA PENDIENTE."


def _sheet(workbook: Workbook, title: str, headers: list[str]):
    if title in workbook.sheetnames:
        del workbook[title]
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def _fit(ws) -> None:
    for column in ws.columns:
        width = min(
            max(len(str(cell.value if cell.value is not None else "")) for cell in column) + 2,
            55,
        )
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def _extracted_documents(state: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(document.get("doc_type")): document.get("extraction") or {}
        for document in state.get("documents", [])
        if document.get("doc_type")
    }


def _cited_value(payload: dict[str, Any], path: str) -> Any:
    current: Any = payload
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    if isinstance(current, dict) and "value" in current:
        return current.get("value")
    return current


def _invoice_summaries(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for line in lines:
        invoice = str(line.get("invoice") or "sin número")
        summary = grouped.setdefault(
            invoice,
            {
                "invoice": invoice,
                "fob": Decimal("0"),
                "freight": Decimal("0"),
                "insurance": Decimal("0"),
                "customs_value": Decimal("0"),
                "landed_cost": Decimal("0"),
                "payable": Decimal("0"),
                "capitalized": Decimal("0"),
                "recoverable": Decimal("0"),
                "levies": {},
                "levy_rates": {},
                "duty_rate": Decimal("0"),
            },
        )
        summary["fob"] += Decimal(str(line.get("fob") or "0"))
        allocations = line.get("allocations") or {}
        summary["freight"] += Decimal(str(allocations.get("freight") or "0"))
        summary["insurance"] += Decimal(str(allocations.get("insurance") or "0"))
        summary["customs_value"] += Decimal(str(line.get("customs_value") or "0"))
        summary["landed_cost"] += Decimal(str(line.get("landed_cost") or "0"))
        declaration_view = line.get("declaration_view") or {}
        cost_view = line.get("cost_view") or {}
        summary["payable"] += Decimal(str(declaration_view.get("payable_levies") or "0"))
        summary["capitalized"] += Decimal(str(cost_view.get("capitalized_levies") or "0"))
        summary["recoverable"] += Decimal(str(cost_view.get("recoverable_levies_excluded") or "0"))
        summary["duty_rate"] = Decimal(str(line.get("duty_rate") or "0"))
        for index, levy in enumerate(line.get("levies") or []):
            summary["levies"][index] = summary["levies"].get(index, Decimal("0")) + Decimal(
                str((levy.get("amount") or {}).get("amount") or "0")
            )
            summary["levy_rates"][str(levy.get("code") or index)] = Decimal(
                str(levy.get("rate") or "0")
            )
    return list(grouped.values())


def _round_control(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _coverage_controls(invoices: list[dict[str, Any]], coverage_pct: Decimal) -> list[Decimal]:
    """Round coverage by invoice while reconciling to the document-level control total."""
    if not invoices:
        return []
    controls = [
        _round_control((invoice["fob"] + invoice["freight"]) * coverage_pct) for invoice in invoices
    ]
    global_control = _round_control(
        sum((invoice["fob"] + invoice["freight"] for invoice in invoices), Decimal("0"))
        * coverage_pct
    )
    residual = global_control - sum(controls, Decimal("0"))
    if residual:
        largest = max(range(len(invoices)), key=lambda index: invoices[index]["fob"])
        controls[largest] += residual
    return controls


def _populate_legacy_summary(workbook: Workbook, invoice_count: int, master_total_row: int) -> None:
    """Extend the operator's legacy summary to the same 100-invoice capacity as the master."""
    ws = workbook["Prorrateo resumen"]
    original_total_row = 32
    total_row = 102
    if ws.max_row < total_row:
        ws.insert_rows(original_total_row, total_row - original_total_row)
        for row in range(original_total_row, total_row):
            for column in range(1, 8):
                source = ws.cell(original_total_row - 1, column)
                target = ws.cell(row, column)
                target._style = copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format

    for row in range(2, total_row):
        has_invoice = row <= invoice_count + 1
        ws.row_dimensions[row].hidden = not has_invoice
        ws[f"A{row}"] = "='Prorrateo General'!A2" if row == 2 and has_invoice else None
        if has_invoice:
            ws[f"B{row}"] = f"='Prorrateo General'!E{row}"
            ws[f"C{row}"] = f"='Prorrateo General'!J{row}"
            ws[f"D{row}"] = f"='Prorrateo General'!F{row}"
            ws[f"E{row}"] = f"='Prorrateo General'!M{row}"
            ws[f"F{row}"] = f"='Prorrateo General'!N{row}"
            ws[f"G{row}"] = f"='Prorrateo General'!C{row}"
        else:
            for column in "BCDEFG":
                ws[f"{column}{row}"] = None

    ws.row_dimensions[total_row].hidden = False
    ws[f"A{total_row}"] = "Totales"
    ws[f"B{total_row}"] = f"='Prorrateo General'!E{master_total_row}"
    ws[f"C{total_row}"] = None
    ws[f"D{total_row}"] = f"='Prorrateo General'!F{master_total_row}"
    ws[f"E{total_row}"] = None
    ws[f"F{total_row}"] = None
    ws[f"G{total_row}"] = f"='Prorrateo General'!C{master_total_row}"
    ws.freeze_panes = "A2"


def _populate_master(workbook: Workbook, state: dict[str, Any]) -> None:
    """Populate a copy of the operator's workbook without altering the source template."""
    ws = workbook["Prorrateo General"]
    calculation = state.get("calculation") or {}
    totals = calculation.get("totals") or {}
    lines = calculation.get("lines") or []
    invoices = _invoice_summaries(lines)
    dispatch = state.get("dispatch") or {}
    extracted = _extracted_documents(state)
    sailing = _cited_value(extracted.get("bill_of_lading", {}), "shipped_on_board_date")

    # Extra columns make the legacy sheet capable of config-driven, per-line treatment.
    ws["O1"] = "Tasa derecho"
    ws["O1"]._style = copy(ws["N1"]._style)
    ws["P1"] = "Tasa impuesto"
    ws["P1"]._style = copy(ws["N1"]._style)
    ws["Q1"] = "Cobertura control"
    ws["Q1"]._style = copy(ws["N1"]._style)
    ws.column_dimensions["O"].width = 13
    ws.column_dimensions["P"].width = 13
    ws.column_dimensions["Q"].width = 17
    ws["A2"] = Decimal(str(totals.get("freight", "0")))
    ws["B2"] = Decimal(str(totals.get("fob", "0")))
    ws["B7"] = max((Decimal(str(line.get("duty_rate", "0"))) for line in lines), default=Decimal(0))
    ws["Q2"] = Decimal(str(calculation.get("policy", {}).get("insurance_coverage_pct", "0")))
    coverage_controls = _coverage_controls(invoices, ws["Q2"].value)

    line_capacity = 100
    original_total_row = 52
    total_row = line_capacity + 2
    if total_row > original_total_row:
        ws.insert_rows(original_total_row, total_row - original_total_row)
        for row in range(original_total_row, total_row):
            for column in range(1, 18):
                source = ws.cell(original_total_row - 1, column)
                target = ws.cell(row, column)
                target._style = copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format

    for row in range(2, total_row):
        ws.row_dimensions[row].hidden = row > len(invoices) + 1
        for column in ("C", "J", "L", "M", "N", "O", "P"):
            ws[f"{column}{row}"] = None
        ws[f"D{row}"] = f"=IFERROR(C{row}/$C${total_row},0)"
        ws[f"E{row}"] = f"=ROUND($A$2*D{row},2)"
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = f"=C{row}+E{row}+F{row}"
        ws[f"H{row}"] = f"=ROUND(G{row}*O{row},2)"
        ws[f"I{row}"] = f"=ROUND((G{row}+H{row})*P{row},2)"
        ws[f"K{row}"] = None

    for index, invoice in enumerate(invoices, start=2):
        if index >= total_row:
            raise ValueError("PRORRATEO MASTER admite un máximo de 100 facturas")
        ws[f"C{index}"] = invoice["fob"]
        ws[f"E{index}"] = invoice["freight"]
        ws[f"F{index}"] = invoice["insurance"]
        ws[f"J{index}"] = dispatch.get("referencia")
        if sailing:
            ws[f"L{index}"] = date.fromisoformat(str(sailing))
        ws[f"M{index}"] = dispatch.get("despacho_no")
        ws[f"N{index}"] = invoice["invoice"]
        ws[f"O{index}"] = invoice["duty_rate"]
        ws[f"P{index}"] = invoice["levy_rates"].get("IVA", Decimal("0"))
        ws[f"K{index}"] = coverage_controls[index - 2]

    for column in ("C", "D", "E", "F", "G", "H", "I", "K"):
        ws[f"{column}{total_row}"] = f"=SUM({column}2:{column}{total_row - 1})"
    ws[f"O{total_row}"] = None
    ws[f"P{total_row}"] = None
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{total_row - 1}"
    ws.sheet_view.showGridLines = False
    ws["F1"] = "Prima póliza cliente"
    ws["K1"] = "Cobertura configurada (control)"
    ws["Q1"] = "Factor cobertura"
    ws["Q2"].number_format = "0.00%"
    _populate_legacy_summary(workbook, len(invoices), total_row)


def build_workbook(state: dict[str, Any], template_path: Path = TEMPLATE_PATH) -> bytes:
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró la plantilla operativa: {template_path}")
    workbook = load_workbook(template_path)
    template_hash = hashlib.sha256(template_path.read_bytes()).hexdigest()
    _populate_master(workbook, state)
    dispatch = state.get("dispatch", {})
    calculation = state.get("calculation") or {}
    totals = calculation.get("totals", {})
    invoices = _invoice_summaries(calculation.get("lines") or [])
    processing = state.get("processing") or {}
    review = state.get("review") or {}

    ws = _sheet(workbook, "Resumen", ["Campo", "Valor"])
    summary = [
        ("Aviso", WORKBOOK_NOTICE),
        ("Despacho", dispatch.get("despacho_no")),
        ("Referencia", dispatch.get("referencia")),
        ("Estado", dispatch.get("status")),
        ("Modo de extracción", processing.get("label")),
        ("Compuerta de revisión", "BLOQUEADA" if review.get("blocked") else "APROBADA"),
        ("Valor aduanero USD", totals.get("customs_value")),
        ("Costo puesto USD", totals.get("landed_cost")),
        ("Tributos USD", totals.get("total_payable")),
        ("Pago estimado CLP", totals.get("total_payable_settlement")),
        ("Tipo de cambio", totals.get("fx_rate")),
        ("Fuente TC", totals.get("fx_source")),
        ("Fecha TC", totals.get("fx_date")),
        ("Plantilla operativa", template_path.name),
        ("SHA-256 plantilla", template_hash),
        ("SHA-256 configuración", dispatch.get("jurisdiction_config_hash")),
        ("SHA-256 perfil cliente", dispatch.get("client_config_hash")),
        ("SHA-256 cálculo", (state.get("calculation_run") or {}).get("input_hash")),
        ("Reglas fiscales", "PROVISIONALES — validar con experto aduanero"),
    ]
    for row in summary:
        ws.append(row)
    ws["B2"].font = Font(color=RED, bold=True)

    ws_docs = _sheet(
        workbook,
        "Documentos",
        [
            "Archivo",
            "Tipo",
            "SHA-256",
            "Páginas",
            "Texto",
            "OCR",
            "Confianza",
            "Estado extracción",
            "Parser",
            "Proveedor",
            "Modelo",
        ],
    )
    for document in state.get("documents", []):
        ws_docs.append(
            [
                document.get("filename"),
                document.get("doc_type"),
                document.get("content_hash"),
                document.get("page_count"),
                document.get("has_text_layer"),
                document.get("ocr_used"),
                document.get("classify_confidence"),
                document.get("extraction_status"),
                document.get("extraction_parser"),
                document.get("extraction_provider"),
                document.get("extraction_model"),
            ]
        )

    ws_ext = _sheet(
        workbook,
        "Extracciones",
        ["Archivo", "Ruta", "Valor", "Procedencia", "Página", "Texto fuente", "Confianza"],
    )
    for document in state.get("documents", []):
        _flatten_cited(ws_ext, document.get("filename"), document.get("extraction") or {})

    ws_rules = _sheet(
        workbook,
        "Validaciones",
        ["ID", "Severidad", "Resultado", "Validación", "Detalle", "Acción", "Impacto financiero"],
    )
    for rule in calculation.get("rules", []):
        ws_rules.append(
            [
                rule.get("id"),
                rule.get("severity"),
                rule.get("status"),
                rule.get("title"),
                rule.get("detail"),
                rule.get("suggested_action"),
                str(rule.get("financial_impact") or ""),
            ]
        )

    ws_alloc = _sheet(
        workbook,
        "Prorrateo",
        [
            "Factura",
            "Descripción",
            "HS",
            "FOB",
            "Participación",
            "Flete",
            "Seguro",
            "Valor aduanero",
            "Tasa",
            "Razón",
            "Ajuste residual",
            "Costo puesto",
        ],
    )
    for line in calculation.get("lines", []):
        allocations = line.get("allocations", {})
        ws_alloc.append(
            [
                line.get("invoice"),
                line.get("description"),
                line.get("hs_code"),
                line.get("fob"),
                line.get("share"),
                allocations.get("freight"),
                allocations.get("insurance"),
                line.get("customs_value"),
                line.get("duty_rate"),
                line.get("duty_reason"),
                ", ".join(line.get("residual_codes") or []),
                line.get("landed_cost"),
            ]
        )

    ws_tax = _sheet(
        workbook,
        "Tributos",
        [
            "Factura",
            "Código",
            "Etiqueta",
            "Base",
            "Expresión",
            "Tasa",
            "Monto",
            "Moneda",
            "Recuperable",
        ],
    )
    for line in calculation.get("lines", []):
        for levy in line.get("levies", []):
            ws_tax.append(
                [
                    line.get("invoice"),
                    levy.get("code"),
                    levy.get("label"),
                    levy.get("base_amount", {}).get("amount"),
                    levy.get("base_expression"),
                    levy.get("rate"),
                    levy.get("amount", {}).get("amount"),
                    levy.get("amount", {}).get("currency"),
                    levy.get("recoverable"),
                ]
            )

    ws_declaration = _sheet(
        workbook,
        "Vista declaración",
        ["Factura", "Valor aduanero", "Tributos pagados", "Moneda"],
    )
    ws_cost = _sheet(
        workbook,
        "Vista costo",
        ["Factura", "Costo puesto", "Tributos capitalizados", "IVA recuperable excluido", "Moneda"],
    )
    for invoice in invoices:
        ws_declaration.append(
            [
                invoice["invoice"],
                invoice["customs_value"],
                invoice["payable"],
                totals.get("currency"),
            ]
        )
        ws_cost.append(
            [
                invoice["invoice"],
                invoice["landed_cost"],
                invoice["capitalized"],
                invoice["recoverable"],
                totals.get("currency"),
            ]
        )

    ws_trace = _sheet(workbook, "Trazabilidad", ["Fecha", "Acción", "Detalle"])
    calculation_run = state.get("calculation_run") or {}
    ws_trace.append(
        [
            calculation_run.get("created_at"),
            "calculation_run",
            f"input_sha256={calculation_run.get('input_hash') or ''}; engine={calculation_run.get('engine_version') or ''}",
        ]
    )
    ws_trace.append([None, "client_config", f"sha256={dispatch.get('client_config_hash') or ''}"])
    ws_trace.append(
        [
            None,
            "jurisdiction_config",
            f"sha256={dispatch.get('jurisdiction_config_hash') or ''}",
        ]
    )
    ws_trace.append([None, "workbook_template", f"{template_path.name}; sha256={template_hash}"])
    ws_trace.append(
        [
            None,
            "extraction_mode",
            f"{processing.get('label') or ''}; providers={','.join(processing.get('providers') or [])}; ocr_reused={processing.get('ocr_reused', 0)}",
        ]
    )
    ws_trace.append(
        [
            None,
            "review_gate",
            f"blocked={bool(review.get('blocked'))}; reasons={review.get('reason_count', 0)}",
        ]
    )
    for event in state.get("audit", []):
        ws_trace.append([event.get("created_at"), event.get("action"), str(event.get("payload"))])

    generated = [
        "Resumen",
        "Documentos",
        "Extracciones",
        "Validaciones",
        "Prorrateo",
        "Tributos",
        "Vista declaración",
        "Vista costo",
        "Trazabilidad",
    ]
    for sheet_name in generated:
        _fit(workbook[sheet_name])
    ws_docs.column_dimensions["F"].width = max(ws_docs.column_dimensions["F"].width, 7)
    workbook.active = workbook.sheetnames.index("Prorrateo General")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _flatten_cited(ws, filename: str, value: Any, path: str = "") -> None:
    if isinstance(value, dict) and {"value", "provenance", "confidence"}.issubset(value):
        ws.append(
            [
                filename,
                path,
                value.get("value"),
                value.get("provenance"),
                value.get("page"),
                value.get("source_text"),
                value.get("confidence"),
            ]
        )
        return
    if isinstance(value, dict):
        for key, child in value.items():
            _flatten_cited(ws, filename, child, f"{path}.{key}".strip("."))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            _flatten_cited(ws, filename, child, f"{path}.{index}")
